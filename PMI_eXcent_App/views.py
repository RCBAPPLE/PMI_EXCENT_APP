from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, render
from django.urls import reverse
from .forms import DiagnosticForm
from xhtml2pdf import pisa
from django.template.loader import get_template
from django.http import JsonResponse
import openpyxl
import os
import datetime
from django.conf import settings

from .models import *

# Permet d'initialiser l'application, lorsque l'url est vide
# Cette fonction nous ramène directement à l'accueil
def page_initial(request):

    lang = 'FR'  # Default language is French

    # Affichage initial
    form = DiagnosticForm(lang=lang)

    # Rendu de la page d'accueil avec le formulaire
    return HttpResponseRedirect(reverse("PMI_eXcent_App:page_accueil", kwargs={"lang": lang}))

# Permet d'afficher la page d'accueil et d'enregistrer les informations du diagnostic, dans la base de données
def page_accueil(request, lang):
    if lang not in ['FR', 'EN']:
        lang = 'FR'  # Langue par défaut : Français

    if request.method == 'POST':
        # Gestion de la soumission du formulaire
        form = DiagnosticForm(request.POST, lang=lang)
        if form.is_valid():
            diagnostic = form.save()  # Sauvegarde de l'objet Diagnostic
            # Redirige vers une page des questions ou affiche un succès
            return render(request, 'PMI_eXcent_App/template_question.html', {
                'Diagnostic_id': diagnostic.id,
                'langue': lang,
                'question': get_object_or_404(Template_Question, partie="Mise en Sécurité", id_question="1"),})

        else:
            # Réafficher la page d'accueil avec les erreurs
            return render(request, 'PMI_eXcent_App/page_accueil.html', {
                'form': form,
                'langue': lang,
            })
    else:
        # Affichage initial
        form = DiagnosticForm(lang=lang)

    # Rendu de la page d'accueil avec le formulaire
    return render(request, 'PMI_eXcent_App/page_accueil.html', {
        'form': form,
        'langue': lang,
    })


# Permet de sauter des parties sur la page question
def saut_partie_page_question(request, lang, diagnostic_id, partie, id_question, last_partie, last_id_question):
    # Récupère l'objet Diagnostic correspondant à l'id donné. Si non trouvé, une erreur 404 est levée.
    diagnostic = get_object_or_404(Diagnostic, pk=diagnostic_id)

    # Récupère la question précédente, en fonction de la partie et de l'ID de la question précédentes. Si non trouvé, une erreur 404 est levée.
    last_question = get_object_or_404(Template_Question, partie=last_partie, id_question=last_id_question)

    # Vérifie si la partie actuelle ou la question actuelle sont différentes de la partie ou de la question précédentes.
    if (partie != last_partie) or (id_question != last_id_question):  
        
        # Si les questions sont sautées, on crée un nouvel enregistrement dans la table Result pour indiquer que la question a été sautée.
        Result.objects.create(Diagnostic=diagnostic, Partie=last_partie, Question=last_question.question_texte_FR, Réponse="Question sautée")

        # Rendu de la page avec la question actuelle et la question précédente.
        return render(request, 'PMI_eXcent_App/template_question.html', {
            'Diagnostic_id': diagnostic_id,    # ID du diagnostic
            'langue': lang,                     # Langue demandée (par ex. 'FR', 'EN')
            'question': get_object_or_404(Template_Question, partie=partie, id_question=id_question),   # Question actuelle à afficher
            'bouton_precedente': last_question,  # Question précédente à afficher
        })

    else: 
        # Si la partie et la question sont les mêmes qu'auparavant, on vérifie si dans l'historique
        # il y a eu d'autres questions .
        last_result = (
            Result.objects
            .filter(Diagnostic=diagnostic)
            .exclude(Partie=last_partie, Question=last_question.question_texte_FR)  # Exclut les résultats précédents de la même question
            .order_by('-id')  # Trie les résultats par ID, du plus récent au plus ancien
            .first()           # Récupère le dernier résultat (le plus récent)
        )
        
        if last_result:
            # Si un résultat précédent existe, on fournit alors la page souhaité avec une possibilité de revenir en arrière.
            return render(request, 'PMI_eXcent_App/template_question.html', {
                'Diagnostic_id': diagnostic_id,
                'langue': lang,
                'question': get_object_or_404(Template_Question, partie=partie, id_question=id_question),
                'bouton_precedente': 'true',  # Signale qu'une question précédente a été répondue
            })
        else:
            # Si aucun résultat précédent n'a été trouvé, on rend la page normalement sans la possibilité d'aller en arrière.
            return render(request, 'PMI_eXcent_App/template_question.html', {
                'Diagnostic_id': diagnostic_id,
                'langue': lang,
                'question': get_object_or_404(Template_Question, partie=partie, id_question=id_question),
            })


# Permet de sauter des parties sur la page bilan et causes
def saut_partie_page_bilan(request, lang, diagnostic_id, partie, id_question):
        # Récupère l'objet Diagnostic correspondant à l'id donné. Si non trouvé, une erreur 404 est levée.
        diagnostic = get_object_or_404(Diagnostic, pk=diagnostic_id)

        # Crée une trace dans l'historique, que l'on sort d'une partie Page de diagnostic 
        Result.objects.create(Diagnostic=diagnostic, Partie="Page de diagnostic", Question="", Réponse="")

        return render(request, 'PMI_eXcent_App/template_question.html', {
                    'Diagnostic_id': diagnostic_id,
                    'langue': lang,
                    'question': get_object_or_404(Template_Question, partie=partie, id_question=id_question),
                    })

# Permet la logique de passage etre les différentes pages du questionnaire
def page_question(request, lang, diagnostic_id, partie, id_question, NBreponse):
    # Si la langue fournie n'est pas "FR" ou "EN", la langue par défaut est définie sur "FR" (Français)
    if lang not in ['FR', 'EN']:
        lang = 'FR'  # Langue par défaut est le français

    # Récupère l'objet Diagnostic à partir de l'ID du diagnostic fourni. Si non trouvé, une erreur 404 est levée.
    diagnostic = get_object_or_404(Diagnostic, pk=diagnostic_id)
    
    # Récupère les données de la question précédente en fonction de la partie et de l'ID de la question.
    # Cela utilise la méthode filter() pour filtrer par "partie" et "id_question", et la méthode values()[0] pour obtenir les données de la première question trouvée.
    previous_question_data = Template_Question.objects.filter(partie=partie, id_question=id_question).values()[0]
    
    # Récupère l'objet de la question précédente à partir de l'ID de la question trouvé.
    previous_question = get_object_or_404(Template_Question, pk=previous_question_data["id"])

    # Crée un enregistrement dans la table "Result" pour enregistrer la réponse à la question précédente.
    # La réponse est déterminée par l'index "NBreponse" (qui commence à 1) sur le type de question de la question précédente.
    Result.objects.create(Diagnostic=diagnostic, Partie=partie, Question=previous_question.question_texte_FR, Réponse=previous_question.typeofquestion[NBreponse - 1])

    # Vérification en fonction du numéro de réponse (NBreponse) :
    # Selon la réponse donnée (1 à 5), la logique détermine la question suivante.
    
    if NBreponse == 1:
        # Si la partie suivante est "Page de diagnostic", on redirige vers la page des résultats.
        if previous_question.partie_question_suivante_1 == "Page de diagnostic":
            return HttpResponseRedirect(reverse("PMI_eXcent_App:results", args=(lang,diagnostic.id,partie,
                                            previous_question.id_question_suivante_1,previous_question.question_suivante_1_texte)))
        else:
            # Sinon, on récupère la prochaine question dans la partie et l'ID spécifiés.
            next_question_data = Template_Question.objects.filter(
                partie=previous_question.partie_question_suivante_1,
                id_question=previous_question.id_question_suivante_1
            ).values()[0]

    elif NBreponse == 2:
        if previous_question.partie_question_suivante_2 == "Page de diagnostic":
            return HttpResponseRedirect(reverse("PMI_eXcent_App:results", args=(lang,diagnostic.id,partie,
                                            previous_question.id_question_suivante_2,previous_question.question_suivante_2_texte)))
        else:
            next_question_data = Template_Question.objects.filter(
                partie=previous_question.partie_question_suivante_2,
                id_question=previous_question.id_question_suivante_2
            ).values()[0]

    elif NBreponse == 3:
        if previous_question.partie_question_suivante_3 == "Page de diagnostic":
            return HttpResponseRedirect(reverse("PMI_eXcent_App:results", args=(lang,diagnostic.id,partie,
                                            previous_question.id_question_suivante_3,previous_question.question_suivante_3_texte)))
        else:
            next_question_data = Template_Question.objects.filter(
                partie=previous_question.partie_question_suivante_3,
                id_question=previous_question.id_question_suivante_3
            ).values()[0]

    elif NBreponse == 4:
        if previous_question.partie_question_suivante_4 == "Page de diagnostic":
            return HttpResponseRedirect(reverse("PMI_eXcent_App:results", args=(lang,diagnostic.id,partie,
                                            previous_question.id_question_suivante_4,previous_question.question_suivante_4_texte)))
        else:
            next_question_data = Template_Question.objects.filter(
                partie=previous_question.partie_question_suivante_4,
                id_question=previous_question.id_question_suivante_4
            ).values()[0]

    elif NBreponse == 5:
        if previous_question.partie_question_suivante_5 == "Page de diagnostic":
            return HttpResponseRedirect(reverse("PMI_eXcent_App:results", args=(lang,diagnostic.id,partie,
                                            previous_question.id_question_suivante_5,previous_question.question_suivante_5_texte)))
        else:
            next_question_data = Template_Question.objects.filter(
                partie=previous_question.partie_question_suivante_5,
                id_question=previous_question.id_question_suivante_5
            ).values()[0]

    # Une fois la prochaine question déterminée, on récupère l'objet correspondant à cette question.
    next_question = get_object_or_404(Template_Question, pk=next_question_data["id"])

    # Rendu du template avec la question suivante et un indicateur pour afficher le bouton précédente.
    return render(request, 'PMI_eXcent_App/template_question.html', {
            'Diagnostic_id': diagnostic.id,  # ID du diagnostic
            'langue': lang,  # Langue sélectionnée (par exemple, "FR" ou "EN")
            'question': next_question,  # Question suivante à afficher
            'bouton_precedente': 'true',  # Indicateur pour afficher le bouton précédente
        })


# permet d'enregistrer le commentaire fait sur la page question
def enregistrer_commentaire(request):
    if request.method == 'POST':
        # Récupération des données
        commentaire = request.POST.get('comment')
        diagnostic_id = request.POST.get('diagnostic_id')
        question_id = request.POST.get('question_id')
        partie = request.POST.get('partie')

        if commentaire:  # Vérifie que le commentaire n'est pas vide
            # Création du commentaire dans la base de données
            Commentaire.objects.create(
                diagnostic_id=diagnostic_id,
                partie=partie,
                id_question=question_id,
                commentaire=commentaire
            )
            return JsonResponse({'success': True})  # Retour de succès

    return JsonResponse({'success': False}, status=400)  # Retour d'échec si pas de données ou problème


# Permet à l'utilisateur de retourner en arrière et modifier ses réponses
def page_precedente(request, lang, diagnostic_id, partie, id_question):
    # Si la langue fournie n'est ni "FR" ni "EN", la langue par défaut est définie sur "FR" (Français)
    if lang not in ['FR', 'EN']:
        lang = 'FR'  # Langue par défaut est le français

    # Récupère l'objet Diagnostic correspondant à l'ID du diagnostic fourni. Si non trouvé, une erreur 404 est levée.
    diagnostic = get_object_or_404(Diagnostic, pk=diagnostic_id)
    
    # Récupère la question spécifiée par la partie et l'ID de la question.
    question_page = get_object_or_404(Template_Question, partie=partie, id_question=id_question)

    # Recherche le dernier résultat enregistré pour le diagnostic, excluant la question actuelle.
    last_result = (
        Result.objects
        .filter(Diagnostic=diagnostic)  # Filtre par diagnostic
        .exclude(Partie=partie, Question=question_page.question_texte_FR)  # Exclut la question actuelle
        .order_by('-id')  # Trie les résultats par ID de manière décroissante (du plus récent au plus ancien)
        .first()  # Récupère le dernier résultat (le plus récent)
    )

    # Si un dernier résultat est trouvé et que la partie n'est pas "Page de diagnostic"
    if last_result and last_result.Partie != "Page de diagnostic":
        last_result.delete()  # Supprime ce dernier résultat pour revenir en arrière
        
        # Trouve la question précédente à partir des informations du dernier résultat (partie et texte de la question)
        previous_question_data = Template_Question.objects.filter(
            partie=last_result.Partie, 
            question_texte_FR=last_result.Question
        ).values()[0]
        previous_question = get_object_or_404(Template_Question, pk=previous_question_data["id"])

        # Recherche si dans l'historique, il y a une autre question avant previous_question
        last_result2 = (
            Result.objects
            .filter(Diagnostic=diagnostic)
            .exclude(Partie=previous_question.partie, Question=previous_question.question_texte_FR)
            .order_by('-id')
            .first()
        )
        
        # S'il existe une autre question enregistrée dans l'historique avant previous_question et que sa partie n'est pas "Page de diagnostic"
        if last_result2 and last_result2.Partie != "Page de diagnostic":
            # Rend la page avec la question précédente et un bouton permettant de revenir plus en arrière dans l'historique
            return render(request, 'PMI_eXcent_App/template_question.html', {
                'Diagnostic_id': diagnostic.id,  # ID du diagnostic
                'langue': lang,  # Langue de la question
                'question': previous_question,  # Question précédente à afficher
                'bouton_precedente': 'true',  # Indicateur pour afficher le bouton "Précédente"
            })
        else:
            # Si aucun résultat précédent valide n'est trouvé, rend la question précédente sans le bouton "Précédente"
            return render(request, 'PMI_eXcent_App/template_question.html', {
                'Diagnostic_id': diagnostic.id,
                'langue': lang,
                'question': previous_question,  # Question précédente
            })

    else:
        # Si aucun résultat précédent n'est trouvé, rend la page de la question actuelle sans bouton "Précédente"
        return render(request, 'PMI_eXcent_App/template_question.html', {
            'Diagnostic_id': diagnostic.id,
            'langue': lang,
            'question': question_page,  # Question actuelle
        })



# Renvoie à la page résultat en affichant le bilan, les causes racines et les moyens de résolution
def page_results(request, lang, diagnostic_id, partie, id_solution, txt_solution):
    # Récupérer le diagnostic
    diagnostic = get_object_or_404(Diagnostic, pk=diagnostic_id)
    

    # Récupérer les résultats associés
    results = Result.objects.filter(Diagnostic=diagnostic)
    
    # Initialiser la solution à None par défaut
    solution = None
    
    # Essayer de récupérer les données des solutions
    try:
        solution_data = Cause.objects.filter(
            Partie=partie,
            Indice=id_solution
        ).values()[0]  # Récupérer la première solution trouvée (si existante)
        
        # Si des données sont trouvées, récupérer l'objet Cause correspondant
        solution = Cause.objects.get(pk=solution_data["id"])
    except (IndexError, Cause.DoesNotExist):
        # Aucune solution trouvée ou erreur d'accès aux données
        pass

    # Rendre le template avec ou sans solution
    return render(request, 'PMI_eXcent_App/bilan_causes.html', {
        'Diagnostic': diagnostic,
        'results': results,
        'langue': lang,
        'Solution': solution,  # Peut être None si aucune solution n'est trouvée
        'Txt_autre': txt_solution, # Au cas où solution pas dans Bdd, affichage de secours
    })

#Création et remplissage du fichier PDF récapitulatif, en se basant sur un modèle
def generate_pdf(request, diagnostic_id, lang):
    # Récupére les données du diagnostic
    diagnostic = get_object_or_404(Diagnostic, pk=diagnostic_id)
    results = Result.objects.filter(Diagnostic=diagnostic)

    # Crée le contexte pour le modèle
    context = {
        'diagnostic': diagnostic,
        'results': results,
        'langue': lang
    }

    # Charge le modèle et le rendre sous forme de chaîne
    template = get_template('PMI_eXcent_App/pdf_template.html')
    html = template.render(context)

    # Créer le PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Diagnostic_{diagnostic.Client}.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)

    # Retourne un fichier PDF ou un message d'erreur
    if pisa_status.err:
        return HttpResponse('Error generating PDF', status=500)
    return response


# Création et modification du fichier RIN
def generate_excel_based_on_language(request, diagnostic_id, lang):
    # Récupére le diagnostic
    diagnostic = Diagnostic.objects.get(id=diagnostic_id)

    # Détermine le template en fonction de la langue
    if lang == 'FR':
        template_path = os.path.join(settings.BASE_DIR, "PMI_eXcent_App/static/docs/template_RIN_fr.xlsx")
    else:
        template_path = os.path.join(settings.BASE_DIR, "PMI_eXcent_App/static/docs/template_RIN_en.xlsx")

    # Charge le template Excel
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active  # Feuille active

    # Rempli les cellules avec les données spécifiques
    ws["S6"] = diagnostic.Client  # Cellule S6 : Nom du client
    ws["S4"] = f"Date: {datetime.datetime.now().strftime('%d/%m/%Y')}"  # Cellule S4 : Date du jour
    ws["R15"] = "JacXson U70"  # Cellule R15 : Texte fixe
    ws["G17"] = diagnostic.SN_JacXson  # Cellule G17 : SN JacXson
    ws["N10"] = datetime.datetime.now().strftime('%d/%m/%Y')

    # Prépare la réponse HTTP avec le fichier Excel
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="RIN_{diagnostic.Client}.xlsx"'
    wb.save(response)  # Sauvegarde le fichier dans la réponse
    return response